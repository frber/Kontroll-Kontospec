import os
import openpyxl
from openpyxl import load_workbook
import re
import numpy as np
import pickle
from openpyxl.styles import Alignment


def appenda_listor_kontospec():
    #Appendar 3 olika listor för att sedan kunna jämföra mot de faktiska filerna och kunna ta fram "nettot" av dessa.
    lista_1270 = []
    lista_1274 = []
    kontospec_lista = []
    wb = openpyxl.load_workbook("kontospec.xlsx")
    for sheet in wb.worksheets:
        konto = sheet.cell(14, 1).value
        if konto == None or konto == "*":
            pass
        else:
            konto = konto[:4]
            # Appendar lista för alla konto 1270. Tar bara projektnummer, ej projektnummer + konto. Jag har valt detta pga av att det finns en variation i hur folk döper sina filer för just mappen 1270. Vissa använder projnummer + konto, andra inte.
            # Har lagt till belopp > 1000 eftersom det finns vissa poster i kontospec för 1270 med några ören/kronor till följd av avrundingsfel vid aktiveringstillfälle. Programmet reagerar annars därför att det finns i 1270, men ej 1274 eller mapp.
            if konto == "1270":
                for row in sheet['C1:C400']:
                    for cell in row:
                        projnummer = cell.value
                        belopp = cell.offset(column=2).value
                        if projnummer == None or isinstance(projnummer, str) or projnummer == 220999 or belopp == None:
                            pass
                        else:
                            if projnummer > 100 and belopp > 1000:
                                lista_1270.append(projnummer)
            # Appendar en lista för alla konto 1274. Ingen beloppsgräns verkar behövas här eftersom det inte finns några sådana poster i 1274. Om det finns i framtiden(?), lägg till beloppsgräns även här.
            if konto == "1274":
                for row in sheet['C1:C400']:
                    for cell in row:
                        projnummer = cell.value
                        if projnummer == None or isinstance(projnummer, str) or projnummer == 220999:
                            pass
                        else:
                            if projnummer > 100:
                                lista_1274.append(projnummer)
            # Appendar alla gjorda periodiseringar (från kontospec), förutom 1270 och 1274, till en lista. Appendar med projektnummer + namn eftersom detta är standarden på hur berpers ska döpas (vilket sedan jämförs mot en lista med faktiska filnamn från mapp).
            else:
                for row in sheet['C1:C400']:
                    for cell in row:
                        projnummer = cell.value
                        if projnummer == None or isinstance(projnummer, str) or projnummer == 220999:
                            pass

                        else:
                            if projnummer > 100 and konto != "1270" and konto != "1274":
                                filnamn = str(projnummer) + str(konto)
                                kontospec_lista.append(filnamn)
    wb.close()
    return  lista_1270, lista_1274, kontospec_lista


def diff_1270(lista_1270, lista_1274, bas_filvag):
    # Rensar filändelse, tar bara projektnummer (fram till första space) från mappen 1270 och lägger in dessa namn i en lista.
    fakt_filer_1270 = []
    filvag_1270 = os.path.join(bas_filvag, "1270. Påg nyanlägg\\")
    for root, dirs, files in os.walk(filvag_1270):
        for file in files:
            fil_rensad = os.path.splitext(file)[0]
            fil_rensad = fil_rensad.rsplit(" ")
            fil_rensad = fil_rensad[0]
            fil_rensad = fil_rensad.replace(".", "")
            fil_rensad = fil_rensad.replace(",", "")
            fil_rensad = fil_rensad.replace("-", "")
            fil_rensad = fil_rensad.replace(" ", "")
            fil_rensad = re.sub('\D', '', fil_rensad)
            fakt_filer_1270.append(int(fil_rensad))

    # Nettot av detta blir filerna som saknas när man jämför mot 1270 kontospec. Alltså: 1270 kontospec - 1274 kontospec - de filer som finns i mappen.
    # Spelar därför ingen roll om det är för mycket filer i mappen eftersom det är kontospecen man rättar sig efter.
    netto = list(set(lista_1270)-set(lista_1274)-set(fakt_filer_1270))
    return netto


def diff_ovriga(kspec_lista, bas_filvag):
    mapp_lista = []
    for root, dirs, files in os.walk(bas_filvag):
        for file in files:
            fil_rensad = os.path.splitext(file)[0]
            fil_rensad = fil_rensad.replace(".", "")
            fil_rensad = fil_rensad.replace(",", "")
            fil_rensad = fil_rensad.replace("-", "")
            fil_rensad = fil_rensad.replace(" ", "")
            fil_rensad = re.sub('\D', '', fil_rensad)
            mapp_lista.append(fil_rensad)

    # Nettot av listan från kontospec (exkl. 1270 och 1274) och listan från alla filer i bokslutsmapp.
    # Nettot är alltså de filer som saknas.
    netto = set(kspec_lista)-set(mapp_lista)
    return netto


def kontrollera_felplacering(bas_filvag):
    lista_felplacering = []
    # Laddar in en lista jag har skapat med ett tidigare program. Den här listan består av två index. Index 0 är kontonummer och index 1 är vilken plats det kontonummret ska vara i e-pärmen.
    lista_kontovag = pickle.load(open('lista_konto_vag','rb'))

    for root, dirs, files in os.walk(bas_filvag):
        for file in files:
            fil_rensad = os.path.splitext(file)[0]
            fil_rensad = fil_rensad.replace(".", "")
            fil_rensad = fil_rensad.replace(",", "")
            fil_rensad = fil_rensad.replace("-", "")
            fil_rensad = fil_rensad.replace(" ", "")
            fil_rensad = re.sub('\D', '', fil_rensad)
            konto = fil_rensad[-4:]

            for x in lista_kontovag:
                if konto == str(x[0]):
                    borde_vara_i = os.path.join(bas_filvag, x[1])
                    ar_i_nu = root
                    if ar_i_nu != borde_vara_i:
                        lista_felplacering.append([fil_rensad, ar_i_nu, borde_vara_i])
    return lista_felplacering


def skriv_utfall(saknas_1270, saknas_ovriga, lista_felplacering):

    wb = openpyxl.load_workbook("Sakande och felplacerade filer.xlsx")
    ws_saknad = wb["Saknade filer"]
    ws_felplacerad = wb["Felplacerade filer"]
    ws_saknad.delete_rows(2,500)
    ws_felplacerad.delete_rows(2,500)

    for y in saknas_1270:
        kst_1270 = str(y)
        kst_1270 = kst_1270[0:3]

        kst1270_ut = ws_saknad.cell(row=ws_saknad.max_row+1, column=1)
        konto1270 = ws_saknad.cell(row=ws_saknad.max_row, column=2)
        filnamn1270 = ws_saknad.cell(row=ws_saknad.max_row, column=3)

        kst1270_ut.value = kst_1270
        kst1270_ut.alignment = Alignment(horizontal='center')
        konto1270.value = 1270
        konto1270.alignment = Alignment(horizontal='center')
        filnamn1270.value = str(y)
        filnamn1270.alignment = Alignment(horizontal='center')

    for x in saknas_ovriga:
        langd = len(x) - 4
        sl = [x[i:i + langd] for i in range(0, len(x), langd)]
        filnamn = ' '.join(sl)
        kst = str(filnamn)
        kst = kst[0:3]
        konto = filnamn[-4:]

        kst_ut = ws_saknad.cell(row=ws_saknad.max_row+1, column=1)
        konto_ut = ws_saknad.cell(row=ws_saknad.max_row, column=2)
        filnamn_ut = ws_saknad.cell(row=ws_saknad.max_row, column=3)

        kst_ut.value = int(kst)
        kst_ut.alignment = Alignment(horizontal='center')
        konto_ut.value = int(konto)
        konto_ut.alignment = Alignment(horizontal='center')
        filnamn_ut.value = filnamn
        filnamn_ut.alignment = Alignment(horizontal='center')

    for z in lista_felplacering:
        kst_felplac = str(z[0])
        kst_felplac = kst_felplac[0:3]
        langd2 = len(z[0]) - 4
        sl2 = [z[0][i:i + langd2] for i in range(0, len(z[0]), langd2)]
        filnamn_felplac = ' '.join(sl2)
        nuvarande_placering = z[1]
        borde_vara_placerad = z[2]

        kst_felplac_ut = ws_felplacerad.cell(row=ws_felplacerad.max_row + 1, column=1)
        filnamn_felplac_ut = ws_felplacerad.cell(row=ws_felplacerad.max_row, column=2)
        nuvarande_placering_ut = ws_felplacerad.cell(row=ws_felplacerad.max_row, column=3)
        borde_vara_placerad_ut = ws_felplacerad.cell(row=ws_felplacerad.max_row, column=4)

        kst_felplac_ut.value = int(kst_felplac)
        kst_felplac_ut.alignment = Alignment(horizontal='center')
        filnamn_felplac_ut.value = filnamn_felplac
        filnamn_felplac_ut.alignment = Alignment(horizontal='center')
        nuvarande_placering_ut.value = nuvarande_placering
        borde_vara_placerad_ut.value = borde_vara_placerad

    wb.save("Sakande och felplacerade filer.xlsx")
    wb.close()


def main():
    bas_filvag = "K:\\TVM\Administration\\3 ADM-GRUPP\\E-BOKSLUT\\2021\\E-pärm 202104\\4. BR, underlag enl innehållsförteckn\\" # Ändras för varje nytt bokslut.
    listor_kspec = appenda_listor_kontospec()
    lista_1270 = listor_kspec[0]
    lista_1274 = listor_kspec[1]
    kspec_lista = listor_kspec[2]
    saknas_1270 = diff_1270(lista_1270, lista_1274, bas_filvag)
    saknas_ovriga = diff_ovriga(kspec_lista, bas_filvag)
    lista_felplacering = kontrollera_felplacering(bas_filvag)
    skriv_utfall(saknas_1270, saknas_ovriga, lista_felplacering)
    os.startfile("Sakande och felplacerade filer.xlsx")


if __name__ == "__main__":
    main()







